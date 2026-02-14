from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table
from openpyxl import load_workbook


def excel_to_pdf(excel_path, pdf_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    data = []

    for r in range(1, ws.max_row + 1):
        row = []
        for c in range(1, ws.max_column + 1):
            row.append(str(ws.cell(row=r, column=c).value or ""))
        data.append(row)

    pdf = SimpleDocTemplate(pdf_path, pagesize=A4)
    table = Table(data)

    pdf.build([table])
