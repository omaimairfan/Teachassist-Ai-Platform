from docx import Document
from openpyxl import load_workbook


def excel_to_word(excel_path, word_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    doc = Document()

    table = doc.add_table(rows=1, cols=ws.max_column)
    table.style = "Table Grid"

    # headers
    for c in range(1, ws.max_column + 1):
        table.rows[0].cells[c - 1].text = str(ws.cell(row=1, column=c).value or "")

    # data
    for r in range(2, ws.max_row + 1):
        row_cells = table.add_row().cells
        for c in range(1, ws.max_column + 1):
            row_cells[c - 1].text = str(ws.cell(row=r, column=c).value or "")

    doc.save(word_path)
