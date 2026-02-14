from openpyxl import load_workbook


def inject_into_template(source_df, mapping, template_path, output_path, output_type):
    if output_type != "xlsx":
        raise Exception("Only Excel supported")

    wb = load_workbook(template_path)
    ws = wb.active

    start_row = list(mapping.values())[0]["row"]

    for i, (_, src_row) in enumerate(source_df.iterrows()):
        excel_row = start_row + i

        for info in mapping.values():
            col = info["column"]
            value = src_row[info["source_column"]]

            cell = ws.cell(row=excel_row, column=col)

            # skip merged cells
            if cell.coordinate in ws.merged_cells:
                continue

            cell.value = value

    wb.save(output_path)
