from openpyxl import load_workbook


def scan_template(template_path: str):
    wb = load_workbook(template_path)
    ws = wb.active

    header_rows = []

    # detect header rows (top 10)
    for r in range(1, 11):
        values = [
            ws.cell(row=r, column=c).value
            for c in range(1, ws.max_column + 1)
            if ws.cell(row=r, column=c).value not in (None, "")
        ]
        if len(values) >= 2:
            header_rows.append(r)

    if not header_rows:
        raise Exception("No header rows found in template")

    data_row = max(header_rows) + 1

    fields = []

    for col in range(1, ws.max_column + 1):
        parts = []
        for r in header_rows:
            cell = ws.cell(row=r, column=col)
            if cell.value:
                parts.append(str(cell.value).strip())

        if parts:
            fields.append({
                "label": " ".join(parts),
                "column": col,
                "row": data_row
            })

    return {
        "fields": fields,
        "data_row": data_row
    }
